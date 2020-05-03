import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
import re
import requests
from datetime import date
from datetime import timedelta
from bs4 import BeautifulSoup
import csv
import pandas as pd
import string

def fixName(s):
    if s == "Japan":
        s = "Japan (+Diamond Princess)"
    if s == "Macau":
        s = "Macao"
    if s == "UK" or s == "U.K.":
        s = "United Kingdom"
    if s == "USA" or s == "U.S.A" or s == "U.S." or s == "US" or s == "USA *":
        s = "United States"
    if s == "UAE" or s == "U.A.E.":
        s = "United Arab Emirates"
    if s == "S. Korea":
        s = "South Korea"
    if s == "Cruise Ship" or s == "Diamond Princess" or s == "Puerto Rico" or s == "U.S. Virgin Islands" or s == "Guam":
        s = "Misc"
    if s == "Czechia" or s == "Czech Republic":
        s = "Czech Republic (Czechia)"
    if s == "Palestine":
        s = "State of Palestine"
    if s == "Vatican City":
        s = "Holy See"
    if s == "St. Barth":
        s = "Saint Barthelemy"
    if s == "DRC":
        s = "DR Congo"
    if s == "Ivory Coast":
        s = "Côte d'Ivoire"
    if s == "St. Vincent Grenadines":
        s = "St. Vincent & Grenadines"
    if s == "CAR":
        s = "Central African Republic"
    if s == "Saint Kitts and Nevis":
        s = "Saint Kitts & Nevis"
    if s == "Saint Pierre Miquelon":
        s = "Saint Pierre & Miquelon"
    if s == "Sao Tome and Principe":
        s = "Sao Tome & Principe"
    return s

def loadTable(url,id):
    page = requests.get(url)
    pagetext = page.text
    soup = BeautifulSoup(pagetext, 'html.parser')
    table = soup.find(id=id)
    return table

def makeString(text):
    if text.strip() == '' or text == "N/A":
        text = "0"
    text = text.replace(',','')
    return text

urls = load_workbook("urls.xlsx",read_only=True)
data = load_workbook("data3.xlsm",keep_vba = True)

invalidVals = ["Country","Total:","World","North America","Europe","Asia","South America","Africa","Oceania",""]

#Create a sheet for each country
page = requests.get("https://www.worldometers.info/coronavirus/countries-where-coronavirus-has-spread/")
pagetext = page.text
soup = BeautifulSoup(pagetext, 'html.parser')
table = soup.find(id="table3")
for row in table.findAll('tr'):
    cells = row.findAll("td")
    if (len(cells)) > 0:
        sheet = data.create_sheet(str(cells[0].text).strip())
        sheet.cell(column=2,row=1).value = "Total cases"
        sheet.cell(column=3,row=1).value = "Total deaths"
        sheet.cell(column=4,row=1).value = "Total recoveries"
        sheet.cell(column=5,row=1).value = "New cases"
        sheet.cell(column=6,row=1).value = "New deaths"
        sheet.cell(column=7,row=1).value = "New recoveries"
        sheet.cell(column=8,row=1).value = "Critical cases"
        sheet.cell(column=9, row=1).value = "Active cases"

data.create_sheet("Misc")

for i in range(2,5):
    url = str(urls["Sheet1"].cell(column=2,row=i).value)
    print(url)
    table = loadTable(url,"table3")
    day = date(2020,1,29) + timedelta(days = i-1)
    for sheet in data.sheetnames[1:]:
        data[sheet].cell(column=1,row=i).value = day.strftime("%Y%m%d")
        data[sheet].cell(column=5,row=i).value = "=B" + str(i) + "-B" + str(i-1)
        data[sheet].cell(column=6, row=i).value = "=C" + str(i) + "-C" + str(i - 1)
        data[sheet].cell(column=7, row=i).value = "=D" + str(i) + "-D" + str(i - 1)
    rows = table.findAll("tr")
    for row in rows:
        cells = row.findAll("td")
        if len(cells) != 0:
            sheet = cells[0].text.strip()
            if sheet != "Country": # not the header
                sheet = fixName(sheet)
                sheet = data[sheet]
                sheet.cell(column=2,row=i).value = int(makeString(cells[1].text))
                sheet.cell(column=3,row=i).value = int(makeString(cells[2].text))
                sheet.cell(column=4,row=i).value = 0

for i in range(5,7):
    url = str(urls["Sheet1"].cell(column=2,row=i).value)
    print(url)
    table = loadTable(url,"table3")
    day = date(2020,1,29) + timedelta(days = i-1)
    for sheet in data.sheetnames[1:]:
        data[sheet].cell(column=1,row=i).value = day.strftime("%Y%m%d")
        data[sheet].cell(column=5, row=i).value = "=B" + str(i) + "-B" + str(i - 1)
        data[sheet].cell(column=6, row=i).value = "=C" + str(i) + "-C" + str(i - 1)
        data[sheet].cell(column=7, row=i).value = "=D" + str(i) + "-D" + str(i - 1)
    rows = table.findAll("tr")
    for row in rows:
        cells = row.findAll("td")
        if len(cells) != 0:
            sheet = cells[0].text.strip()
            if sheet != "Country": # not the header
                sheet = fixName(sheet)
                sheet = data[sheet]
                sheet.cell(column=2,row=i).value = int(makeString(cells[1].text))
                sheet.cell(column=3,row=i).value = int(makeString(cells[3].text))
                sheet.cell(column=4,row=i).value = 0

for i in range(7, 32):
    url = str(urls["Sheet1"].cell(column=2, row=i).value)
    print(url)
    table = loadTable(url, "table3")
    day = date(2020, 1, 29) + timedelta(days=i - 1)
    for sheet in data.sheetnames[1:]:
        data[sheet].cell(column=1, row=i).value = day.strftime("%Y%m%d")
        data[sheet].cell(column=5, row=i).value = "=B" + str(i) + "-B" + str(i - 1)
        data[sheet].cell(column=6, row=i).value = "=C" + str(i) + "-C" + str(i - 1)
        data[sheet].cell(column=7, row=i).value = "=D" + str(i) + "-D" + str(i - 1)
    rows = table.findAll("tr")
    for row in rows:
        cells = row.findAll("td")
        if len(cells) != 0:
            sheet = cells[0].text.strip()
            if sheet != "Country":  # not the header
                sheet = fixName(sheet)
                sheet = data[sheet]
                sheet.cell(column=2, row=i).value = int(makeString(cells[1].text))
                sheet.cell(column=3, row=i).value = int(makeString(cells[3].text))
                sheet.cell(column=4, row=i).value = int(makeString(cells[5].text))
                sheet.cell(column=8, row=i).value = int(makeString(cells[6].text))

for i in range(32, 40):
    url = str(urls["Sheet1"].cell(column=2, row=i).value)
    print(url)
    table = loadTable(url, "main_table_countries")
    day = date(2020, 1, 29) + timedelta(days=i - 1)
    for sheet in data.sheetnames[1:]:
        data[sheet].cell(column=1, row=i).value = day.strftime("%Y%m%d")
        data[sheet].cell(column=5, row=i).value = "=B" + str(i) + "-B" + str(i - 1)
        data[sheet].cell(column=6, row=i).value = "=C" + str(i) + "-C" + str(i - 1)
        data[sheet].cell(column=7, row=i).value = "=D" + str(i) + "-D" + str(i - 1)
    rows = table.findAll("tr")
    for row in rows:
        cells = row.findAll("td")
        if len(cells) != 0:
            sheet = cells[0].text.strip()
            if sheet != "Country" and sheet != "Total:":  # not the header
                sheet = fixName(sheet)
                sheet = data[sheet]
                sheet.cell(column=2, row=i).value = int(makeString(cells[1].text))
                sheet.cell(column=3, row=i).value = int(makeString(cells[3].text))
                sheet.cell(column=4, row=i).value = int(makeString(cells[6].text))
                sheet.cell(column=8, row=i).value = int(makeString(cells[7].text))

for i in range(40, 51):
    url = str(urls["Sheet1"].cell(column=2, row=i).value)
    print(url)
    table = loadTable(url, "main_table_countries")
    day = date(2020, 1, 29) + timedelta(days=i - 1)
    for sheet in data.sheetnames[1:]:
        data[sheet].cell(column=1, row=i).value = day.strftime("%Y%m%d")
        data[sheet].cell(column=5, row=i).value = "=B" + str(i) + "-B" + str(i - 1)
        data[sheet].cell(column=6, row=i).value = "=C" + str(i) + "-C" + str(i - 1)
        data[sheet].cell(column=7, row=i).value = "=D" + str(i) + "-D" + str(i - 1)
    rows = table.findAll("tr")
    for row in rows:
        cells = row.findAll("td")
        if len(cells) != 0:
            sheet = cells[0].text.strip()
            if sheet != "Country" and sheet != "Total:":  # not the header
                sheet = fixName(sheet)
                sheet = data[sheet]
                sheet.cell(column=2, row=i).value = int(makeString(cells[1].text))
                sheet.cell(column=3, row=i).value = int(makeString(cells[3].text))
                sheet.cell(column=4, row=i).value = int(makeString(cells[5].text))
                sheet.cell(column=8, row=i).value = int(makeString(cells[7].text))

i = 51
while not urls["Sheet1"].cell(column=2,row=i).value is None:
    url = str(urls["Sheet1"].cell(column=2, row=i).value)
    print(url)
    table = loadTable(url, "main_table_countries_yesterday")
    day = date(2020, 1, 29) + timedelta(days=i - 1)
    for sheet in data.sheetnames[1:]:
        data[sheet].cell(column=1, row=i).value = day.strftime("%Y%m%d")
        data[sheet].cell(column=5, row=i).value = "=B" + str(i) + "-B" + str(i - 1)
        data[sheet].cell(column=6, row=i).value = "=C" + str(i) + "-C" + str(i - 1)
        data[sheet].cell(column=7, row=i).value = "=D" + str(i) + "-D" + str(i - 1)
    rows = table.findAll("tr")
    for row in rows:
        cells = row.findAll("td")
        if len(cells) != 0:
            sheet = cells[0].text.strip()
            if not sheet in invalidVals:  # not the header
                sheet = fixName(sheet)
                sheet = data[sheet]
                sheet.cell(column=2, row=i).value = int(makeString(cells[1].text))
                sheet.cell(column=3, row=i).value = int(makeString(cells[3].text))
                sheet.cell(column=4, row=i).value = int(makeString(cells[5].text))
                sheet.cell(column=8, row=i).value = int(makeString(cells[7].text))
    i += 1


data.save("data5.xlsm")


